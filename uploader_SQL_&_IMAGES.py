import mysql.connector
from datetime import datetime
from mysql.connector import Error
from openpyxl import load_workbook
import uuid
from PIL import Image
import os
import paramiko
from scp import SCPClient

FOLDER_EXCEL = "D:/DOWN/MiroMark/miromark.xlsx"
FOLDER_PRODUCT = "D:/DOWN/MiroMark/Спальні"
REMOTH_PATH = 'mebli-local.com.ua/www/storage/app/public/img/'

host = ('xi400102.ftp.tools', 'xi400102', '2!bT~i3S')
connect_database = ('xi400102_db', 'gUMq7g9G', 'xi400102.mysql.tools', 'xi400102_db')
# connect_database = ("root", "root", "127.0.0.1", "mebli-site")

list_images = []
image_mass = []
count = 0
brand_id = 3
category_id = 5
status = 1


def main():
    global list_images, image_mass, count
    create_connection(connect_database[0], connect_database[1], connect_database[2], connect_database[3])
    excel = get_excel
    count_folder = len(os.listdir(FOLDER_PRODUCT))
    while count < count_folder:
        post_value_text(excel)
        post_image_url()
        upload_image()
        count += 1
        list_images = []
        image_mass = []
    connection.close()
    print('## Done :-)')


def create_connection(user_name, user_password, host_name, dbase):
    global connection
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            passwd=user_password,
            database=dbase
        )
        print("Connection to MySQL DB successful")
    except Error as e:
        print(f"Error '{e}' occurred")


def execute_query(row, value):
    cursor = connection.cursor()
    try:
        cursor.executemany(row, value)
        id_product = cursor.lastrowid
        connection.commit()
        print(f"Query executed successfully -- {str(value)}")
        return id_product
    except Error as e:
        print(f"The error {e} occurred")


def get_excel(a):
    try:
        wb = load_workbook(FOLDER_EXCEL)
        ws = wb['Sheet1']
        s = ws[a].value
        print("Connected to Excel")
        return s
    except Error as e:
        print(f"Can`t connect to Excel: {e}")


def post_value_text(excel):
    global product_id
    current_Date = datetime.now()
    formatted_date = current_Date.strftime('%Y-%m-%d %H:%M:%S')
    create_product = "INSERT INTO products (name, description, category_id, brand_id, status, created_at, updated_at) " \
                     "VALUES (%s, %s, %s, %s, %s, %s, %s)"
    value_product = [(
        excel('B' + str(count + 1)), excel('C' + str(count + 1)), category_id, brand_id, status, formatted_date,
        formatted_date)]
    product_id = execute_query(create_product, value_product)
    print(product_id)
    sort_image()


def post_image_url():
    current_date = datetime.now()
    formatted_date = current_date.strftime('%Y-%m-%d %H:%M:%S')
    create_image = "INSERT INTO images (product_id, image, size, created_at, updated_at) VALUES (%s, %s, %s, %s, %s)"
    for image in image_mass:
        value_image = [(product_id, "img/{}.jpeg".format(image[0]), str(image[1]) + 'x' + str(image[2]), formatted_date,
                        formatted_date)]
        execute_query(create_image, value_image)


def reformat_image(im_number, path, images_folder):
    size = 1200, 900
    unique_name = im_number + str(uuid.uuid4())
    images = '{}/{}/{}'.format(FOLDER_PRODUCT, path[count], images_folder)
    image = Image.open(images)
    print(f"#### Convert from: {images_folder} - RESOLUTION - {str(image.size)}")
    image.thumbnail(size)
    image = image.convert('RGB')
    save_dir = '{}/{}/{}'.format(FOLDER_PRODUCT, path[count], unique_name)
    image.save(f'{save_dir}.jpeg', 'webp')
    os.remove(images)
    list_images.append(f"{save_dir}.jpeg")
    print(f'## to: {unique_name}.jpeg - NEW RESOLUTION - {str(image.size)}')
    image_mass.append((unique_name, image.size[0], image.size[1]))
    image_mass.sort()


def sort_image():
    path = os.listdir(FOLDER_PRODUCT)
    path.sort(key=int)
    image_path = os.listdir(FOLDER_PRODUCT + "/" + path[count])
    for images_folder in image_path:
        split_image = images_folder.split("_")
        if split_image[-1] != 'medium.jpg' and 'medium.png':
            number = f'0_{split_image[0]}_'
            reformat_image(number, path, images_folder)
        else:
            number = f'1_{split_image[0]}_'
            reformat_image(number, path, images_folder)


def upload_image():
    number = 0
    client = paramiko.SSHClient()
    client.load_host_keys(os.path.join(os.path.dirname(__file__), 'known_hosts'))
    client.connect(host[0], username=host[1], password=host[2])
    scp = SCPClient(client.get_transport())
    for single_image in list_images:
        number += 1
        print(f'## Uploaded -{number} images - {single_image} - to server')
        scp.put(single_image, REMOTH_PATH)
    scp.close()


main()
